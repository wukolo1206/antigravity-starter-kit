import openpyxl
import os
import sys

# 強制控制台輸出為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

folder = os.path.dirname(os.path.abspath(__file__))

def anonymize_name(name):
    if not name:
        return name
    name = str(name).strip()
    n_len = len(name)
    if n_len <= 1:
        return name
    elif n_len == 2:
        return name[0] + '○'
    else:
        return name[0] + '○' * (n_len - 2) + name[-1]

def main():
    resp_path = os.path.join(folder, "學生習作成績登錄表（國習12課、數習9單元） (回覆).xlsx")
    chi_blank_path = os.path.join(folder, "三年2班國語_平時成績匯出入檔_空白.xlsx")
    math_blank_path = os.path.join(folder, "三年2班數學_平時成績匯出入檔_空白.xlsx")
    
    chi_output_path = os.path.join(folder, "三年2班國語_平時成績匯出入檔_已填寫.xlsx")
    math_output_path = os.path.join(folder, "三年2班數學_平時成績匯出入檔_已填寫.xlsx")
    
    # 1. 讀取回覆檔
    print("正在讀取回覆檔...")
    wb_resp = openpyxl.load_workbook(resp_path, data_only=True)
    ws_resp = wb_resp.active
    
    # 解析回覆檔標頭
    header = [cell.value for cell in ws_resp[1]]
    
    # 找出各科目/課堂對應的欄位索引 (0-based)
    col_seat_name = -1
    chi_fields = {} # 課名 -> column index
    math_fields = {} # 單元名 -> column index
    
    for idx, col_name in enumerate(header):
        if not col_name:
            continue
        col_name = str(col_name).strip()
        if "座號與姓名" in col_name:
            col_seat_name = idx
        elif "國語習作" in col_name:
            # 轉換為空白模板的格式，例如將 "國語習作 第 1 課成績" 轉為 "國語習作第1課"
            # 移除空格及"成績"兩個字
            clean_name = col_name.replace(" ", "").replace("成績", "")
            chi_fields[clean_name] = idx
        elif "數學習作" in col_name:
            # 將 "數學習作 第 1 單元成績" 轉為 "數學習作第1單元"
            clean_name = col_name.replace(" ", "").replace("成績", "")
            math_fields[clean_name] = idx

    print(f"解析到座號姓名欄位索引: {col_seat_name}")
    print(f"國語欄位對應: {chi_fields}")
    print(f"數學欄位對應: {math_fields}")
    
    # 讀取學生成績資料 (後填寫的會覆蓋先填寫的)
    scores_dict = {}
    for r in range(2, ws_resp.max_row + 1):
        row_vals = [cell.value for cell in ws_resp[r]]
        if all(v is None for v in row_vals):
            continue
        
        seat_name_val = row_vals[col_seat_name]
        if not seat_name_val:
            continue
            
        parts = str(seat_name_val).strip().split()
        if len(parts) < 1:
            continue
            
        try:
            seat_no = int(parts[0])
        except ValueError:
            print(f"警告: 列 {r} 的座號無法解析: {parts[0]}")
            continue
            
        real_name = parts[1] if len(parts) > 1 else ""
        
        # 國語成績
        student_chi = {}
        for clean_name, col_idx in chi_fields.items():
            if col_idx < len(row_vals):
                student_chi[clean_name] = row_vals[col_idx]
                
        # 數學成績
        student_math = {}
        for clean_name, col_idx in math_fields.items():
            if col_idx < len(row_vals):
                student_math[clean_name] = row_vals[col_idx]
                
        scores_dict[seat_no] = {
            'real_name': real_name,
            'chinese': student_chi,
            'math': student_math
        }
        
    print(f"成功解析了 {len(scores_dict)} 位學生的回覆資料。")
    
    # 2. 處理國語空白檔
    print("\n處理國語科成績與姓名去識別化...")
    wb_chi = openpyxl.load_workbook(chi_blank_path)
    ws_chi = wb_chi.active
    
    # 讀取國語空白檔第 5 列的欄位名稱 (對應科目/課堂)
    chi_cols_map = {} # 課名 -> column index (1-based for openpyxl)
    for col_idx in range(4, ws_chi.max_column + 1):
        val = ws_chi.cell(row=5, column=col_idx).value
        if val:
            clean_val = str(val).strip().replace(" ", "")
            chi_cols_map[clean_val] = col_idx
            
    print(f"國語空白檔科目對應欄位: {chi_cols_map}")
    
    # 填寫國語資料
    r = 7
    filled_chi_count = 0
    while True:
        seat_val = ws_chi.cell(row=r, column=1).value
        if seat_val is None:
            break
        try:
            seat_no = int(seat_val)
        except ValueError:
            r += 1
            continue
            
        # 姓名去識別化
        orig_name = ws_chi.cell(row=r, column=2).value
        anon_name = anonymize_name(orig_name)
        ws_chi.cell(row=r, column=2).value = anon_name
        
        # 填入成績
        if seat_no in scores_dict:
            student_data = scores_dict[seat_no]
            student_scores = student_data['chinese']
            for name_in_sheet, col_idx in chi_cols_map.items():
                score = student_scores.get(name_in_sheet)
                if score is not None:
                    ws_chi.cell(row=r, column=col_idx).value = score
            filled_chi_count += 1
            
        r += 1
        
    wb_chi.save(chi_output_path)
    print(f"國語檔寫入完成，成功填寫了 {filled_chi_count} 位學生的成績，另存於: {chi_output_path}")
    
    # 3. 處理數學空白檔
    print("\n處理數學科成績與姓名去識別化...")
    wb_math = openpyxl.load_workbook(math_blank_path)
    ws_math = wb_math.active
    
    # 讀取數學空白檔第 5 列的欄位名稱 (對應單元)
    math_cols_map = {} # 單元名 -> column index (1-based)
    for col_idx in range(4, ws_math.max_column + 1):
        val = ws_math.cell(row=5, column=col_idx).value
        if val:
            clean_val = str(val).strip().replace(" ", "")
            math_cols_map[clean_val] = col_idx
            
    print(f"數學空白檔單元對應欄位: {math_cols_map}")
    
    # 填寫數學資料
    r = 7
    filled_math_count = 0
    while True:
        seat_val = ws_math.cell(row=r, column=1).value
        if seat_val is None:
            break
        try:
            seat_no = int(seat_val)
        except ValueError:
            r += 1
            continue
            
        # 姓名去識別化
        orig_name = ws_math.cell(row=r, column=2).value
        anon_name = anonymize_name(orig_name)
        ws_math.cell(row=r, column=2).value = anon_name
        
        # 填入成績
        if seat_no in scores_dict:
            student_data = scores_dict[seat_no]
            student_scores = student_data['math']
            for name_in_sheet, col_idx in math_cols_map.items():
                score = student_scores.get(name_in_sheet)
                if score is not None:
                    ws_math.cell(row=r, column=col_idx).value = score
            filled_math_count += 1
            
        r += 1
        
    wb_math.save(math_output_path)
    print(f"數學檔寫入完成，成功填寫了 {filled_math_count} 位學生的成績，另存於: {math_output_path}")

if __name__ == "__main__":
    main()
